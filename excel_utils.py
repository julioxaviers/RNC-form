
import openpyxl
from openpyxl.drawing.image import Image
import datetime

# Get the current date
today = datetime.date.today()

def save_to_excel(data, uploaded_files):
    # Open the Excel file
    wb = openpyxl.load_workbook('C:/Users/JCSILVA/Desktop/RNC/FERRAMENTA RNC 2024-2.xlsx')

    # Select the 'FORMULÁRIO' sheet
    ws = wb['FORMULÁRIO']

    # Adicionar dados à planilha
    
    ws['AI9'] = f"{data['NC']}"
    ws['K13'] = f"{data['Area']}"
    ws['AI13'] = f"{data['Numpal']}"
    ws['K19'] = f"{data['Fornecedor']}"
    ws['V24'] = f"{data['Transportador']}"
    ws['AR24'] = f"{data['Placa']}"
    ws['AG24'] = f"{data['Motorista']}"
    ws['U26'] = f"{data['Tipo_de_NC']}"
    ws['A27'] = f"{data['Descricao']}"
    ws['K21'] = f"{data['Emitente']}"
    ws['AI17'] = f"{data['Turno']}"
    ws['K17'] = f"{data['NF']}"
    #ws['A14'] = f"Qualidade acionada: {data['Qualidade']}"

    # Adicionar imagens à planilha
    image_positions = ['A34', 'L34', 'Z34', 'AJ34']
    for i, file in enumerate(uploaded_files):
        if i >= len(image_positions):
            break  # Apenas as primeiras 4 imagens
        img = Image(file.path)

        # Definir o tamanho exato da imagem (em pixels)
        img.width = 310  # Largura da imagem em pixels
        img.height = 690  # Altura da imagem em pixels

        # Colocar a imagem na célula correta
        cell_position = image_positions[i]
        ws.add_image(img, cell_position)
        

    # Salvar o arquivo Excel
    # Save the Excel file
    wb.save('C:/Users/JCSILVA/Desktop/RNC/RNC '+ str(today) +'-'+ data['NF'] + '.xlsx')
